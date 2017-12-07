# coding=utf-8

"""
Labeling_Tool
Format xlsx to txt.

Written by JNing Wei
"""

src = './Origin'    # dir for origin pics
dst = './Formated'    # dir for formated pics

import os
xlsx_list = [os.path.join(src, name) for name in os.listdir(src) if name.endswith('xlsx')]
xlsx_list.sort()

# remove dst dir every time
import shutil
try:
    shutil.rmtree(dst)
except OSError:
    pass
os.makedirs(dst)

# main operation
from openpyxl import load_workbook
for xlsx_file in xlsx_list:
    r_file = load_workbook(xlsx_file)
    w_file_name = xlsx_file.replace(src, dst).replace("xlsx", "txt")
    w_file = open(w_file_name, "w")
    # print(file.sheetnames)
    sheet = r_file.get_sheet_by_name("Sheet1")
    assert sheet.max_row - 1 == len(sheet["M"][1:])
    w_file.writelines("{}\n".format(str(sheet.max_row - 1)))
    for idx, (xmin, ymin, xmax, ymax) in enumerate(zip(sheet["M"][1:], sheet["N"][1:], sheet["O"][1:], sheet["P"][1:])):
        # print("{:>3d}: ({:>4d},{:>4d}) ({:>4d},{:>4d})".format(idx, xmin.value, ymin.value, xmax.value, ymax.value))
        line = "{0} {1} {2} {3} {4}\n".format(idx, ymin.value, xmin.value, ymax.value, xmax.value)
        w_file.writelines(line)
    w_file.close()

"""
  ┌───┐   ┌───┬───┬───┬───┐ ┌───┬───┬───┬───┐ ┌───┬───┬───┬───┐ ┌───┬───┬───┐
  │Esc│   │ F1│ F2│ F3│ F4│ │ F5│ F6│ F7│ F8│ │ F9│F10│F11│F12│ │P/S│S L│P/B│  ┌┐    ┌┐    ┌┐
  └───┘   └───┴───┴───┴───┘ └───┴───┴───┴───┘ └───┴───┴───┴───┘ └───┴───┴───┘  └┘    └┘    └┘
  ┌───┬───┬───┬───┬───┬───┬───┬───┬───┬───┬───┬───┬───┬───────┐ ┌───┬───┬───┐ ┌───┬───┬───┬───┐
  │~ `│! 1│@ 2│# 3│$ 4│% 5│^ 6│& 7│* 8│( 9│) 0│_ -│+ =│ BacSp │ │Ins│Hom│PUp│ │N L│ / │ * │ - │
  ├───┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─────┤ ├───┼───┼───┤ ├───┼───┼───┼───┤
  │ Tab │ Q │ W │ E │ R │ T │ Y │ U │ I │ O │ P │{ [│} ]│ | \ │ │Del│End│PDn│ │ 7 │ 8 │ 9 │   │
  ├─────┴┬──┴┬──┴┬──┴┬──┴┬──┴┬──┴┬──┴┬──┴┬──┴┬──┴┬──┴┬──┴─────┤ └───┴───┴───┘ ├───┼───┼───┤ + │
  │ Caps │ A │ S │ D │ F │ G │ H │ J │ K │ L │: ;│" '│ Enter  │               │ 4 │ 5 │ 6 │   │
  ├──────┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴─┬─┴────────┤     ┌───┐     ├───┼───┼───┼───┤
  │ Shift  │ Z │ X │ C │ V │ B │ N │ M │< ,│> .│? /│  Shift   │     │ ↑ │     │ 1 │ 2 │ 3 │   │
  ├─────┬──┴─┬─┴──┬┴───┴───┴───┴───┴───┴──┬┴───┼───┴┬────┬────┤ ┌───┼───┼───┐ ├───┴───┼───┤ E││
  │ Ctrl│    │Alt │         Space         │ Alt│    │    │Ctrl│ │ ← │ ↓ │ → │ │   0   │ . │←─┘│
  └─────┴────┴────┴───────────────────────┴────┴────┴────┴────┘ └───┴───┴───┘ └───────┴───┴───┘

  Code is far away from bug with the keyboard protecting.
"""
