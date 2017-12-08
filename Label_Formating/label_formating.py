# coding=utf-8

"""
Labeling_Tool
Format xlsx to txt.

__author__ = 'JNingWei'
"""

src = "./Origin"    # dir for origin pics
dst = "./Formated"    # dir for formated pics

origin_suffix = "xlsx"
new_suffix = "txt"

import os
xlsx_list = [os.path.join(src, name) for name in os.listdir(src) if name.endswith(origin_suffix)]
xlsx_list.sort()

# remove dst dir every time
import shutil
try:
    shutil.rmtree(dst)
except OSError:
    pass
os.makedirs(dst)

# main operation
from openpyxl import load_workbook
for xlsx_file in xlsx_list:
    r_file = load_workbook(xlsx_file)
    w_file_name = xlsx_file.replace(src, dst).replace(origin_suffix, new_suffix)
    w_file = open(w_file_name, "w")
    # print(file.sheetnames)
    sheet = r_file.get_sheet_by_name("Sheet1")
    assert sheet.max_row - 1 == len(sheet["M"][1:])
    w_file.writelines("{}\n".format(str(sheet.max_row - 1)))
    for (xmin, ymin, xmax, ymax) in zip(sheet["M"][1:], sheet["N"][1:], sheet["O"][1:], sheet["P"][1:]):
        # print("{:>3d}: ({:>4d},{:>4d}) ({:>4d},{:>4d})".format(idx, xmin.value, ymin.value, xmax.value, ymax.value))
        line = "{0} {1} {2} {3}\n".format(xmin.value, ymin.value, xmax.value, ymax.value)
        w_file.writelines(line)
    w_file.close()


